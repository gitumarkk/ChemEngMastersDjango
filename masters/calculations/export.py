from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

class ExportToExcel(object):
    def __init__(self, data):
        self.data = data
        self.wb = Workbook()

    def write_summary(self):
        # The summary will always be the first worksheet
        ws = self.wb.active
        ws.title = "Reactor Summary"
        summary_data = self.data["summary"]

        _r = 1

        for k, v in summary_data.iteritems():
            if k == "data":
                continue
            _c = 1
            col = get_column_letter(_c)
            ws.cell('%s%s'%(col, _r)).value = k
            _r += 1
            for k_2, v_2 in v.iteritems():
                if k_2 == 'initial_component_conc' or k_2 == 'final_component_conc':
                    c_3 = 1
                    ws.cell('%s%s'%("A", _r)).value = k_2
                    for k_3, v_3 in v_2.iteritems():
                        col = get_column_letter(c_3)
                        ws.cell('%s%s'%(col, _r+1)).value = k_3
                        ws.cell('%s%s'%(col, _r+2)).value = v_3  # _r + 1 so that data can be in bottom
                        c_3 += 1
                    _r += 5
                else:
                    col = get_column_letter(_c)
                    ws.cell('%s%s'%(col, _r)).value = k_2
                    ws.cell('%s%s'%(col, _r+1)).value = v_2  # _r + 1 so that data can be in bottom
                _c += 1
            _r += 3  # Leaving a space and accomodating for _r+1 above

    def write_general_data(self, ws, data):
        _r = 2
        map_data_col = {"step": "A"}
        _char = "B"
        for row in data:
            col = get_column_letter(1)  # Assuming 1st entry is for Time
            if _r == 2: # First pas through the loop need to update the header
                ws.cell('%s%s'%(col, 1)).value = "Time (Min)"

                # BUILD HASH MAPS OF EXCEL COLUMNS TO VALUES AS DICT DOES NOT MAINTAIN POSITION
                for k in row["ions"].iterkeys():
                    map_data_col[k] = _char
                    _char = chr(ord(_char) + 1)

            ws.cell('%s%s'%(col, _r)).value = row["step"]

            _c = 2
            for k, v in row["ions"].iteritems():
                # col = get_column_letter(_c) # Reserver first column for step
                if _r == 2: # First pass through the loop need to update the header
                    ws.cell('%s%s'%(map_data_col[k], 1)).value = k
                ws.cell('%s%s'%(map_data_col[k], _r)).value = v

                # ws.cell('%s%s'%(col, 1)).value = k
                # ws.cell('%s%s'%(col, _r)).value = v
                # if k == "Sn":
                #     print _c, ": ",col, ": ",_r, ": ",v, ": ",len(row["ions"])
            # if _r > 360:
            #     print row["ions"]
            #     # print k, v
                _c += 1
            _r += 1

    def write_bioxidation(self):
        biox_ws = self.wb.create_sheet()
        biox_ws.title = "Bioxidation Data"

        biox_data = self.data["bioxidation"]
        self.write_general_data(biox_ws, biox_data)

    def write_chemical(self):
        chem_ws = self.wb.create_sheet()
        chem_ws.title = "Chemical Data"

        chem_data = self.data["chemical"]
        self.write_general_data(chem_ws, chem_data)

    def run(self):
        self.write_summary()
        self.write_bioxidation()
        self.write_chemical()
        # Export to Django
        return save_virtual_workbook(self.wb)
